"""
ERP Parser (Group D)
====================

Parser for Damodaran Country Risk Premium data.
Outputs: erp.json (178 countries)

Data source: ctryprem.xlsx
- ERPs by country sheet: country, region, rating, default_spread, erp, crp
- Country Tax Rates sheet: country, corporate_tax_rate

Version: 1.0.0 (2026-01-11)
Note: Migrated from damodaran-erp/fetcher.py with 2026 structure support
"""

import json
import logging
import requests
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Optional

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import URLS, OUTPUT_FILES
from base.percent_parser import parse_percent

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

# openpyxl required for xlsx files
try:
    import openpyxl
except ImportError:
    openpyxl = None


class ERPParser:
    """
    ERP (Equity Risk Premium) parser for country-based data.

    Output structure (erp.json):
    {
        "metadata": {...},
        "countries": {
            "United States": {
                "equity_risk_premium": 0.0446,
                "country_risk_premium": 0.0023,
                "default_spread": 0.0,
                "region": "North America",
                "rating": "Aa1",
                "corporate_tax_rate": 0.25
            }
        },
        "us_erp": 0.0446
    }
    """

    # 2026 structure: ERPs by country sheet
    # Typical 2026 structure. The parser verifies the header row dynamically.
    SHEET_NAME = "ERPs by country"
    HEADER_ROW = 8
    DATA_START_ROW = 9

    # Column mapping for 2026 structure
    COL_MAP = {
        "country": 1,
        "region": 2,
        "rating": 3,        # Moody's rating
        "default_spread": 4,
        "erp": 5,           # Total Equity Risk Premium
        "crp": 6,           # Country Risk Premium
    }

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize parser.

        Args:
            output_dir: Output directory for JSON files
        """
        self.output_dir = output_dir or Path(__file__).parent.parent / "output"
        self.countries: Dict[str, Dict[str, Any]] = {}
        self.us_erp: Optional[float] = None

    def download(self, timeout: int = 60) -> bytes:
        """
        Download ERP Excel file.

        Args:
            timeout: Request timeout in seconds

        Returns:
            Raw content bytes
        """
        url = URLS.get("erp")
        if not url:
            raise ValueError("ERP URL not configured in config.py")

        logger.info(f"Downloading ERP data from {url}...")
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()

        logger.info(f"  Downloaded {len(response.content) / 1024:.1f} KB")
        return response.content

    def parse(self, content: bytes) -> Dict[str, Dict[str, Any]]:
        """
        Parse ERP data from Excel content.

        Args:
            content: Raw Excel file bytes

        Returns:
            Dictionary of country data
        """
        if openpyxl is None:
            raise ImportError("openpyxl required: pip install openpyxl")

        logger.info("Parsing ERP data...")
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

        # ERPs by country sheet (2026 structure)
        ws = wb[self.SHEET_NAME]
        logger.info(f"  Sheet: {ws.title}")

        header_row = self._find_header_row(ws)
        data_start_row = header_row + 1
        col_map = self._build_column_map(ws, header_row)
        logger.info(f"  Header row: {header_row}")
        logger.info(f"  Column mapping: {col_map}")

        # Parse country data
        countries = {}
        for row_idx in range(data_start_row, ws.max_row + 1):
            country = ws.cell(row=row_idx, column=col_map["country"]).value
            if not country:
                continue

            country_str = str(country).strip()

            # Skip invalid values
            if country_str in ["", "#N/A", "N/A"]:
                continue

            # Parse ERP value
            erp_val = ws.cell(row=row_idx, column=col_map["erp"]).value
            erp = parse_percent(erp_val)

            if erp is None:
                continue

            # Parse other fields
            region = ws.cell(row=row_idx, column=col_map["region"]).value if "region" in col_map else None
            region_str = str(region).strip() if region else None

            rating = ws.cell(row=row_idx, column=col_map["rating"]).value if "rating" in col_map else None
            rating_str = str(rating).strip() if rating else None

            default_spread = parse_percent(
                ws.cell(row=row_idx, column=col_map["default_spread"]).value
            ) if "default_spread" in col_map else None
            crp = parse_percent(
                ws.cell(row=row_idx, column=col_map["crp"]).value
            ) if "crp" in col_map else None

            countries[country_str] = {
                "equity_risk_premium": erp,
                "country_risk_premium": crp,
                "default_spread": default_spread,
                "region": region_str,
                "rating": rating_str
            }

            # Track US ERP
            if country_str.lower() in ["united states", "us", "usa", "u.s."]:
                self.us_erp = erp

        # Merge tax rates from separate sheet
        tax_rates = self._parse_tax_rates(wb)
        merged_count = 0
        for country_name, tax_rate in tax_rates.items():
            if country_name in countries:
                countries[country_name]["corporate_tax_rate"] = tax_rate
                merged_count += 1

        self.countries = countries
        logger.info(f"  Parsed {len(countries)} countries (Tax Rate merged: {merged_count})")

        if self.us_erp:
            logger.info(f"  US Equity Risk Premium: {self.us_erp:.2%}")

        return countries

    def _find_header_row(self, ws) -> int:
        """Find the ERP country header row dynamically."""
        for row_idx in range(1, min(20, ws.max_row + 1)):
            values = [
                str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                for col_idx in range(1, min(12, ws.max_column + 1))
            ]
            joined = " ".join(values)
            has_country_col = any(value == "country" for value in values)
            has_erp_col = "total equity risk premium" in joined or any(
                value == "equity risk premium" for value in values
            )
            if has_country_col and has_erp_col:
                return row_idx
        logger.warning("ERP header row not detected dynamically; using configured default")
        return self.HEADER_ROW

    def _build_column_map(self, ws, header_row: int) -> Dict[str, int]:
        """Build column map from the ERP header row."""
        col_map: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if not cell_val:
                continue
            col_str = str(cell_val).lower().replace(" ", "").replace("/", "")
            if "country" == col_str or "country" in col_str and "risk" not in col_str:
                col_map["country"] = col_idx
            elif "region" in col_str:
                col_map["region"] = col_idx
            elif col_idx == 2 and "region" not in col_map:
                # April 2026 file labels this column "Africa" even though it
                # contains region names for all countries.
                col_map["region"] = col_idx
            elif "defaultspread" in col_str:
                col_map["default_spread"] = col_idx
            elif "mood" in col_str:
                col_map["rating"] = col_idx
            elif "rating" in col_str and "rating" not in col_map:
                col_map["rating"] = col_idx
            elif "totalequityriskpremium" in col_str and "erp" not in col_map:
                col_map["erp"] = col_idx
            elif "equityriskpremium" in col_str and "country" not in col_str and "erp" not in col_map:
                col_map["erp"] = col_idx
            elif "countryriskpremium" in col_str and "crp" not in col_map:
                col_map["crp"] = col_idx

        if "country" not in col_map or "erp" not in col_map:
            logger.warning("ERP dynamic column map incomplete; using configured defaults")
            return self.COL_MAP.copy()
        return col_map

    def _parse_tax_rates(self, wb) -> Dict[str, float]:
        """
        Parse Country Tax Rates sheet.

        Args:
            wb: openpyxl workbook

        Returns:
            Dict[country_name, tax_rate]
        """
        tax_rates = {}
        try:
            ws = wb["Country Tax Rates"]
            logger.info(f"  Tax Rate sheet: {ws.title}")

            # Row 1: Header ("Country", "Tax Rate", ...)
            # Row 2+: Data
            for row_idx in range(2, ws.max_row + 1):
                country = ws.cell(row=row_idx, column=1).value
                tax_val = ws.cell(row=row_idx, column=2).value

                if not country:
                    continue

                country_str = str(country).strip()
                if country_str in ["", "#N/A", "N/A"]:
                    continue

                tax_rate = parse_percent(tax_val)
                if tax_rate is not None:
                    tax_rates[country_str] = tax_rate

            logger.info(f"  Tax Rate parsed: {len(tax_rates)} countries")
        except KeyError:
            logger.warning("Country Tax Rates sheet not found - Tax Rate omitted")

        return tax_rates

    def to_json(self, output_path: Optional[Path] = None, us_only: bool = False) -> Path:
        """
        Save parsed data to JSON.

        Args:
            output_path: Output file path (default: output/erp.json)
            us_only: If True, only save US data

        Returns:
            Path to saved file
        """
        if not self.countries:
            raise ValueError("No data parsed. Call parse() first.")

        # Determine output path
        if output_path is None:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            output_path = self.output_dir / OUTPUT_FILES["erp"]
        else:
            output_path = Path(output_path)

        # Build JSON structure
        result = {
            "metadata": {
                "source": "Damodaran Online",
                "url": URLS.get("erp", ""),
                "schema_version": "2.0.0",
                "generated_at": datetime.now().isoformat(),
                "source_date": "April 1, 2026",
                "scope": "Country-level equity risk premiums",
                "source_format": "Damodaran XLSX current data file",
                "country_count": len(self.countries) if not us_only else 1,
                "note": "2026 structure: ERPs by country sheet + Tax Rate merged"
            }
        }

        if us_only:
            # US only
            us_data = None
            for country, data in self.countries.items():
                if country.lower() in ["united states", "us", "usa", "u.s."]:
                    us_data = {"United States": data}
                    break
            result["countries"] = us_data or {}
        else:
            result["countries"] = self.countries

        result["us_erp"] = self.us_erp

        # Save
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)

        file_size = output_path.stat().st_size / 1024
        logger.info(f"Saved to: {output_path} ({file_size:.1f} KB)")

        return output_path

    def convert(self, output_path: Optional[Path] = None, us_only: bool = False) -> Path:
        """
        Full conversion pipeline: download -> parse -> save.

        Args:
            output_path: Output path (default: output/erp.json)
            us_only: If True, only save US data

        Returns:
            Path to saved file
        """
        content = self.download()
        self.parse(content)
        return self.to_json(output_path, us_only)

    def get_us_erp(self) -> Optional[float]:
        """
        Get US Equity Risk Premium.

        Returns:
            US ERP as decimal (e.g., 0.0446 for 4.46%)
        """
        if not self.countries:
            content = self.download()
            self.parse(content)
        return self.us_erp


# =============================================================================
# CLI
# =============================================================================

def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(description="Parse Damodaran ERP data")
    parser.add_argument(
        "-o", "--output",
        help="Output file path (default: output/erp.json)"
    )
    parser.add_argument(
        "--us-only",
        action="store_true",
        help="Only output US data"
    )
    args = parser.parse_args()

    # Parse
    erp = ERPParser()
    output_path = erp.convert(
        output_path=Path(args.output) if args.output else None,
        us_only=args.us_only
    )

    # Summary
    print("\n" + "=" * 60)
    print(f"Countries: {len(erp.countries)}")
    print(f"US ERP: {erp.us_erp:.2%}" if erp.us_erp else "US ERP: N/A")
    print(f"Output: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
