"""
Excel Loader
============

Unified Excel file loader supporting both .xls (xlrd) and .xlsx (openpyxl) formats.
"""

import logging
from io import BytesIO
from typing import Any, List, Optional, Tuple

# Optional import for xlrd (used for .xls files)
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    xlrd = None
    XLRD_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelWorksheet:
    """
    Unified worksheet interface for both xlrd and openpyxl.

    Provides consistent cell access regardless of underlying library.
    """

    def __init__(self, ws, library: str):
        """
        Initialize worksheet wrapper.

        Args:
            ws: Underlying worksheet object (xlrd or openpyxl)
            library: "xlrd" or "openpyxl"
        """
        self._ws = ws
        self._library = library

    @property
    def title(self) -> str:
        """Get sheet name."""
        if self._library == "xlrd":
            return self._ws.name
        return self._ws.title

    @property
    def max_row(self) -> int:
        """Get maximum row number."""
        if self._library == "xlrd":
            return self._ws.nrows
        return self._ws.max_row

    @property
    def max_column(self) -> int:
        """Get maximum column number."""
        if self._library == "xlrd":
            return self._ws.ncols
        return self._ws.max_column

    def cell(self, row: int, column: int) -> Any:
        """
        Get cell value.

        Args:
            row: Row number (1-indexed for consistency)
            column: Column number (1-indexed)

        Returns:
            Cell value
        """
        if self._library == "xlrd":
            # xlrd uses 0-indexed rows and columns
            try:
                return self._ws.cell_value(row - 1, column - 1)
            except IndexError:
                return None
        else:
            # openpyxl uses 1-indexed
            return self._ws.cell(row=row, column=column).value


class ExcelWorkbook:
    """
    Unified workbook interface for both xlrd and openpyxl.
    """

    def __init__(self, content: bytes):
        """
        Load Excel file from bytes content.

        Args:
            content: Raw bytes of Excel file

        Raises:
            ImportError: If required library not installed
            ValueError: If file format not recognized
        """
        self._workbook = None
        self._library = None
        self._sheets = {}

        # Try openpyxl first (xlsx)
        try:
            import openpyxl
            self._workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
            self._library = "openpyxl"
            logger.debug("Loaded as xlsx with openpyxl")
            return
        except Exception as e:
            logger.debug(f"openpyxl failed: {e}")

        # Try xlrd (xls)
        try:
            import xlrd
            self._workbook = xlrd.open_workbook(file_contents=content)
            self._library = "xlrd"
            logger.debug("Loaded as xls with xlrd")
            return
        except ImportError:
            raise ImportError("xlrd required for .xls files: pip install xlrd")
        except Exception as e:
            logger.debug(f"xlrd failed: {e}")
            raise ValueError(f"Could not load Excel file: {e}")

    @property
    def sheetnames(self) -> List[str]:
        """Get list of sheet names."""
        if self._library == "xlrd":
            return self._workbook.sheet_names()
        return self._workbook.sheetnames

    @property
    def active(self) -> ExcelWorksheet:
        """Get active (first) sheet."""
        if self._library == "xlrd":
            ws = self._workbook.sheet_by_index(0)
        else:
            ws = self._workbook.active
        return ExcelWorksheet(ws, self._library)

    def __getitem__(self, name: str) -> ExcelWorksheet:
        """
        Get sheet by name.

        Args:
            name: Sheet name

        Returns:
            ExcelWorksheet wrapper

        Raises:
            KeyError: If sheet not found
        """
        if name in self._sheets:
            return self._sheets[name]

        if self._library == "xlrd":
            try:
                ws = self._workbook.sheet_by_name(name)
            except xlrd.XLRDError:
                raise KeyError(f"Sheet '{name}' not found")
        else:
            if name not in self._workbook.sheetnames:
                raise KeyError(f"Sheet '{name}' not found")
            ws = self._workbook[name]

        wrapper = ExcelWorksheet(ws, self._library)
        self._sheets[name] = wrapper
        return wrapper

    def __contains__(self, name: str) -> bool:
        """Check if sheet exists."""
        return name in self.sheetnames


def load_excel(content: bytes) -> ExcelWorkbook:
    """
    Load Excel file from bytes content.

    Automatically detects format (xls vs xlsx) and uses appropriate library.

    Args:
        content: Raw bytes of Excel file

    Returns:
        ExcelWorkbook instance

    Example:
        >>> from base.excel_loader import load_excel
        >>> wb = load_excel(response.content)
        >>> ws = wb["Sheet1"]
        >>> value = ws.cell(row=1, column=1)
    """
    return ExcelWorkbook(content)
