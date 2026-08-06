#!/usr/bin/env python3
# FENO RIM v2 — ERP archive restoration (reproducible from committed raw bytes).
#
# Parses the Damodaran country-risk-premium archive files
# (data/damodaran/archives/ctryprem{YY}.xls[x], fetched 2026-08-06 with raw-byte
# receipts in receipt.json) into dated US/Korea composite-ERP observations, then
# builds the W=52 weekly-state band per the amended contract (section 6).
#
# FIRST-KNOWABLE DATE RULE (owner ruling 2026-08-06, uniform):
#   base = the official archive-page label, ctryprem{YY} -> {YY+1}-01-31
#          (end of the stated month; conservative within the month);
#   if an internal date is present AND LATER than the base -> use the internal
#          date (more conservative; ctryprem15's 2016-07-01 is a genuine
#          mid-year edition);
#   if an internal date is present and EARLIER than the base -> keep the base
#          and record the rejected value with the reason (stale cells:
#          ctryprem22's 2018-12-31, ctryprem24's 2020-12-31).
#   No observation may carry month_unknown — the archive page supplies it.
#   `date_source` records which level produced each date.
#
# COLUMN RULE: US and Korea read the file's composite country ERP (the Total
# Equity Risk Premium column where a header exists, structural fallback
# otherwise) — never additive US+CRP, per the contract.

import xlrd
import openpyxl
import glob
import os
import re
import json
import datetime
import hashlib

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT = os.path.join(ROOT, "data", "damodaran", "archives")
ARTIFACT = os.path.join(ROOT, "data", "computed", "feno-rim-v2", "erp-archive-restoration.json")
EVALUATION_DATE = datetime.date(2026, 8, 6)
EXCEL_EPOCH = datetime.date(1899, 12, 30)
MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def serial_to_date(v):
    try:
        return EXCEL_EPOCH + datetime.timedelta(days=int(float(v)))
    except Exception:
        return None


def load_sheets(path):
    if path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return {n: [list(r) for r in wb[n].iter_rows(values_only=True)] for n in wb.sheetnames}
    wb = xlrd.open_workbook(path)
    return {n: [[wb.sheet_by_name(n).cell_value(r, c) for c in range(wb.sheet_by_name(n).ncols)]
                for r in range(wb.sheet_by_name(n).nrows)] for n in wb.sheet_names()}


def find_us_kr(rows):
    us = kr = None
    for i, row in enumerate(rows):
        for cell in row[:3]:
            s = str(cell).strip()
            if us is None and (re.match(r"^united states", s, re.I) or re.match(r"^u\.?s\.?$", s, re.I)):
                us = (i, row)
            if kr is None and re.match(r"^korea\b", s, re.I) and not re.search(r"north|d\.?p\.?r", s, re.I):
                kr = (i, row)
        if us and kr:
            break
    return us, kr


def find_erp_column(rows, us_row_idx, us_row):
    for i in range(min(us_row_idx, len(rows)) - 1, -1, -1):
        cells = [str(c) for c in rows[i]]
        if not any("country" in c.lower() for c in cells):
            continue
        for j, c in enumerate(cells):
            low = c.lower()
            if "total equity risk" in low or "total erp" in low:
                return j, f"header:{c.strip()[:28]}"
        for j, c in enumerate(cells):
            low = c.lower()
            if re.match(r"^erp$", low) or "composite erp" in low or "equity risk premium" in low or "risk premium" in low:
                return j, f"header:{c.strip()[:28]}"
    for j in range(2, len(us_row)):
        try:
            v = float(us_row[j])
            if 0.01 < v < 0.35:
                return j, "structural:first_plausible_us_value"
        except Exception:
            pass
    return None, "header_not_found"


def internal_vintage(sheets):
    """(date, note) from the file itself; None when absent."""
    for name, rows in sheets.items():
        if name.lower() in ("raw ratings", "regional breakdown", "regional simple averages",
                            "relative equity volatility", "regional lookup table", "country lookup",
                            "sheet2", "sheet3"):
            continue
        for row in rows[:16]:
            for cell in row:
                m = re.search(r"updated\s*(?::)?\s*([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})", str(cell), re.I)
                if m:
                    month = MONTHS.get(m.group(1).lower())
                    if month:
                        return datetime.date(int(m.group(3)), month, int(m.group(2))), f"updated_text:{str(cell).strip()[:40]}"
        for row in rows[:16]:
            for j, cell in enumerate(row):
                if re.search(r"date of update", str(cell), re.I) and j + 1 < len(row):
                    v = row[j + 1]
                    d = v if isinstance(v, datetime.datetime) else serial_to_date(v)
                    if isinstance(d, datetime.datetime):
                        d = d.date()
                    if d:
                        return d, f"date_of_update_cell:{v!r}"
    return None, "no_internal_date"


def apply_date_rule(internal, yy):
    """Uniform first-knowable rule (owner ruling 2026-08-06)."""
    base = datetime.date(2000 + yy + 1, 1, 31)
    if internal is None:
        return base, "page_label", None
    if internal > base:
        return internal, "internal_date_later", None
    return base, "page_label_internal_rejected", internal


def parse_file(path):
    tag = os.path.basename(path)
    yy = int(re.search(r"ctryprem(\d{2})", tag).group(1))
    sheets = load_sheets(path)
    data_name = next((n for n in sheets if n.lower() == "erps by country"),
                     next((n for n in sheets if n.lower() == "country premiums"),
                          next((n for n in sheets if n.lower().startswith("sheet1")), None)))
    if data_name is None:
        return {"file": tag, "error": "no_data_sheet"}
    rows = sheets[data_name]
    us, kr = find_us_kr(rows)
    if not us or not kr:
        return {"file": tag, "error": f"us_row={us is not None} kr_row={kr is not None}"}
    col, col_note = find_erp_column(rows, us[0], us[1])
    if col is None:
        return {"file": tag, "error": f"erp_column:{col_note}"}
    try:
        us_erp = float(us[1][col])
        kr_erp = float(kr[1][col])
    except Exception as e:
        return {"file": tag, "error": f"value_parse:{e}"}
    if not (0.01 < us_erp < 0.35 and 0.01 < kr_erp < 0.35):
        return {"file": tag, "error": f"sanity us={us_erp} kr={kr_erp} col={col}"}
    internal, v_note = internal_vintage(sheets)
    first_knowable, date_source, rejected = apply_date_rule(internal, yy)
    raw = open(path, "rb").read()
    return {
        "file": tag,
        "sheet": data_name,
        "yy": yy,
        "edition_page_label": f"{yy + 1:02d}-01-31",
        "internal_date": internal.isoformat() if internal else None,
        "internal_date_note": v_note,
        "first_knowable": first_knowable.isoformat(),
        "date_source": date_source,
        "rejected_internal_date": rejected.isoformat() if rejected else None,
        "rejected_reason": (f"internal_date_earlier_than_page_month_end" if rejected else None),
        "us_erp": round(us_erp, 6),
        "kr_erp": round(kr_erp, 6),
        "erp_column": col,
        "erp_column_note": col_note,
        "us_row": us[0] + 1,
        "kr_row": kr[0] + 1,
        "byte_size": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
    }


def weekly_states(observations, end):
    ok = [o for o in observations if "error" not in o]
    ok.sort(key=lambda o: o["first_knowable"])
    start = datetime.date.fromisoformat(ok[0]["first_knowable"])
    week = start - datetime.timedelta(days=start.weekday())
    us_states, kr_states = [], []
    us_rel = kr_rel = None
    while week <= end:
        for o in ok:
            fk = datetime.date.fromisoformat(o["first_knowable"])
            if fk <= week:
                us_rel = kr_rel = o
        us_states.append({"week": week.isoformat(), "erp": us_rel["us_erp"] if us_rel else None,
                          "release_file": us_rel["file"] if us_rel else None})
        kr_states.append({"week": week.isoformat(), "erp": kr_rel["kr_erp"] if kr_rel else None,
                          "release_file": kr_rel["file"] if kr_rel else None})
        week += datetime.timedelta(days=7)
    return us_states, kr_states


def band_at(us_states, kr_states, end):
    """52 point-in-time weekly states ending at `end`: min/max + distinct releases."""
    win_us = [s for s in us_states if s["week"] <= end.isoformat()][-52:]
    win_kr = [s for s in kr_states if s["week"] <= end.isoformat()][-52:]
    us_vals = [s["erp"] for s in win_us if s["erp"] is not None]
    kr_vals = [s["erp"] for s in win_kr if s["erp"] is not None]
    us_rel = sorted({s["release_file"] for s in win_us if s["release_file"]})
    kr_rel = sorted({s["release_file"] for s in win_kr if s["release_file"]})
    return {
        "evaluation_date": end.isoformat(),
        "us": {"low": round(min(us_vals), 6), "high": round(max(us_vals), 6),
               "distinct_official_releases": len(us_rel), "releases": us_rel,
               "point": min(us_vals) == max(us_vals)},
        "kr": {"low": round(min(kr_vals), 6), "high": round(max(kr_vals), 6),
               "distinct_official_releases": len(kr_rel), "releases": kr_rel,
               "point": min(kr_vals) == max(kr_vals)},
    }


def origin_grid():
    """Quarterly origins over the walk-forward span (panel start 2010-01-01)."""
    d = datetime.date(2010, 1, 1)
    out = []
    while d <= datetime.date(2023, 5, 14):
        out.append(d)
        d += datetime.timedelta(days=91)
    return out


def build_artifact():
    observations = [parse_file(p) for p in sorted(glob.glob(os.path.join(OUT, "ctryprem*.xls*")))]
    ok = [o for o in observations if "error" not in o]
    us_states, kr_states = weekly_states(ok, EVALUATION_DATE)
    named = [datetime.date(y, 6, 30) for y in (2015, 2018, 2021, 2024)] + [EVALUATION_DATE]
    named_bands = [band_at(us_states, kr_states, d) for d in named]
    origin_bands = [band_at(us_states, kr_states, d) for d in origin_grid()]
    points = sum(1 for b in origin_bands if b["us"]["point"] and b["kr"]["point"])
    body = {
        "schema_version": "fenok_rim_erp_archive_restoration.v1",
        "status": "RESTORED",
        "generated_at": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "raw_archive_dir": OUT,
        "raw_archive_receipt": "receipt.json",
        "observations": observations,
        "observations_ok": len(ok),
        "observations_failed": [o for o in observations if "error" in o],
        "weekly_states_us": us_states,
        "weekly_states_kr": kr_states,
        "band_us_52w_at_evaluation": band_at(us_states, kr_states, EVALUATION_DATE)["us"],
        "band_kr_52w_at_evaluation": band_at(us_states, kr_states, EVALUATION_DATE)["kr"],
        "named_evaluation_bands": named_bands,
        "origin_band_summary": {
            "origins": len(origin_bands),
            "point_windows": points,
            "range_windows": len(origin_bands) - points,
            "distinct_release_distribution": {
                str(n): sum(1 for b in origin_bands if b["us"]["distinct_official_releases"] == n)
                for n in sorted({b["us"]["distinct_official_releases"] for b in origin_bands})
            },
        },
        "date_rule": "page_label {YY+1}-01-31 base; internal later -> internal; internal earlier -> base + rejected recorded",
        "notes": [
            "ctryprem25 (Jan 2026 edition) absent from the archive - reported, not worked around",
            "point windows are the honest consequence of an annual release under a 52-week state window; the distinct-release count is published to expose exactly this",
        ],
    }
    body_sha_input = {k: v for k, v in body.items() if k != "generated_at"}
    body_sha = hashlib.sha256(json.dumps(body_sha_input, sort_keys=True).encode()).hexdigest()
    body["artifact_sha256"] = body_sha
    os.makedirs(os.path.dirname(ARTIFACT), exist_ok=True)
    with open(ARTIFACT, "w") as f:
        json.dump(body, f, indent=2)
    return body


if __name__ == "__main__":
    artifact = build_artifact()
    for o in artifact["observations"]:
        if "error" in o:
            print(o["file"], "ERROR", o["error"])
        else:
            print(f"{o['file']}: fk={o['first_knowable']} src={o['date_source']} "
                  f"rejected={o['rejected_internal_date'] or '-'} us={o['us_erp']} kr={o['kr_erp']}")
    print("\nband at evaluation date:", artifact["band_us_52w_at_evaluation"], artifact["band_kr_52w_at_evaluation"])
    for b in artifact["named_evaluation_bands"]:
        print(f"  {b['evaluation_date']}: US {b['us']['low']}-{b['us']['high']} distinct={b['us']['distinct_official_releases']} "
              f"| KR {b['kr']['low']}-{b['kr']['high']} distinct={b['kr']['distinct_official_releases']}")
    print("origin summary:", artifact["origin_band_summary"])
    print(f"sha256 {artifact['artifact_sha256'][:16]}… written {ARTIFACT}")
